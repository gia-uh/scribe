from __future__ import annotations

import io

from openpyxl import load_workbook

from ..result import ExtractResult


def _rows_to_md(rows: list[list]) -> str:
    rows = [
        [("" if c is None else str(c)) for c in r]
        for r in rows
        if any(c is not None for c in r)
    ]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    head, *body = rows
    out = ["| " + " | ".join(head) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for r in body:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def convert(data: bytes) -> ExtractResult:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"could not open XLSX: {exc}") from exc

    parts: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        md = _rows_to_md(rows)
        if md:
            parts.append(f"### {name}\n\n{md}")
    return ExtractResult(
        markdown="\n\n".join(parts), meta={"backend": "xlsx", "sheets": len(wb.sheetnames)}
    )
